import customtkinter as ctk
import tkinter.filedialog as fd
import threading
import cv2
import pytesseract
import os
import re
import subprocess
import openpyxl
import time
import shutil 
import sys
import datetime
from PIL import Image, ImageDraw, ImageOps, ImageTk

# ==========================================
# PORTABLE COMPILATION PATH ENGINE
# ==========================================
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    bundle_dir = sys._MEIPASS
    possible_tesseract_path = os.path.join(bundle_dir, "tesseract")
    if not os.path.exists(possible_tesseract_path):
        possible_tesseract_path = os.path.join(bundle_dir, "Contents", "Resources", "tesseract")
        
    pytesseract.pytesseract.tesseract_cmd = possible_tesseract_path
    os.environ["TESSDATA_PREFIX"] = os.path.join(bundle_dir, "tessdata")
else:
    pytesseract.pytesseract.tesseract_cmd = r'/opt/homebrew/bin/tesseract'

# ==========================================
# UI THEME SETTINGS
# ==========================================
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class JobTile(ctk.CTkFrame):
    def __init__(self, master, video_name, **kwargs):
        super().__init__(master, fg_color="#2b2b2b", corner_radius=8, **kwargs)
        
        self.grid_columnconfigure((0, 1, 2, 3), weight=1)
        self.start_time = None
        self.is_running = False
        
        self.lbl_title = ctk.CTkLabel(self, text=video_name, font=ctk.CTkFont(size=14, weight="bold"), anchor="w")
        self.lbl_title.grid(row=0, column=0, columnspan=4, padx=15, pady=(10, 0), sticky="w")
        
        self.lbl_status = ctk.CTkLabel(self, text="Waiting in queue...", text_color="#a3a3a3", font=ctk.CTkFont(size=12), anchor="w")
        self.lbl_status.grid(row=1, column=0, columnspan=4, padx=15, pady=0, sticky="w")
        
        self.progressbar = ctk.CTkProgressBar(self, height=8, corner_radius=4)
        self.progressbar.grid(row=2, column=0, columnspan=4, padx=15, pady=(5, 10), sticky="ew")
        self.progressbar.set(0)

        self.lbl_duration = ctk.CTkLabel(self, text="Duration: --:--", font=ctk.CTkFont(size=11), text_color="#a3a3a3")
        self.lbl_duration.grid(row=3, column=0, padx=15, pady=(0, 10), sticky="w")

        self.lbl_elapsed = ctk.CTkLabel(self, text="Elapsed: 00:00", font=ctk.CTkFont(size=11), text_color="#a3a3a3")
        self.lbl_elapsed.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="w")

        self.lbl_stage = ctk.CTkLabel(self, text="Stage: Queued", font=ctk.CTkFont(size=11, weight="bold"), text_color="#3B8ED0")
        self.lbl_stage.grid(row=3, column=2, padx=5, pady=(0, 10), sticky="w")

        self.lbl_resized = ctk.CTkLabel(self, text="Resized: Pending", font=ctk.CTkFont(size=11), text_color="#a3a3a3")
        self.lbl_resized.grid(row=3, column=3, padx=15, pady=(0, 10), sticky="e")

    def set_duration(self, seconds):
        mins, secs = divmod(int(seconds), 60)
        self.lbl_duration.configure(text=f"Duration: {mins:02d}:{secs:02d}")

    def start_timer(self):
        self.start_time = time.time()
        self.is_running = True
        self.update_clock()

    def stop_timer(self):
        self.is_running = False

    def update_clock(self):
        if not self.is_running: return
        elapsed = int(time.time() - self.start_time)
        mins, secs = divmod(elapsed, 60)
        self.lbl_elapsed.configure(text=f"Elapsed: {mins:02d}:{secs:02d}")
        self.after(1000, self.update_clock)

    def update_tile(self, status_text, progress_val=None, color="#ffffff", stage=None, resized_status=None):
        self.lbl_status.configure(text=status_text, text_color=color)
        if progress_val is not None: self.progressbar.set(progress_val)
        if stage is not None: self.lbl_stage.configure(text=f"Stage: {stage}")
        if resized_status is not None: self.lbl_resized.configure(text=f"Resized: {resized_status}")


class VideoResizerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Video Resizer")
        self.geometry("1050x700")
        
        icon_path = "app_icon.png"
        if os.path.exists(icon_path):
            icon_img = ImageTk.PhotoImage(Image.open(icon_path))
            self.iconphoto(True, icon_img)
            
        self.excel_path = None
        self.video_folder = None
        self.bg_path = None
        self.output_video_folder = None
        self.col_indices = {} 
        
        self.user_choice = None
        self.wait_event = threading.Event()
        self.cancel_flag = False 

        # ==========================================
        # GLOBAL SETTINGS DICTIONARY
        # ==========================================
        self.cfg = {
            "shrink": 0.87,
            "bottom_pct": 0.13,
            "corner_pct": 0.09,
            "crf": 28,
            "preset": "slower",
            "time_buffer_start": 0.5,
            "max_bridge_gap": 15.0  # UPDATED: Holds shrink for 15 seconds
        }

        # --- GRID LAYOUT ---
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- SIDEBAR ---
        self.sidebar_frame = ctk.CTkFrame(self, width=330, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(8, weight=1) 

        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="Video Resizer", font=ctk.CTkFont(size=26, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(30, 20))

        btn_params = {"width": 270, "height": 45, "font": ctk.CTkFont(size=13, weight="bold"), "corner_radius": 8}

        self.btn_select_excel = ctk.CTkButton(self.sidebar_frame, text="1. Select Existing Excel", command=self.select_excel, **btn_params)
        self.btn_select_excel.grid(row=1, column=0, padx=20, pady=10)

        self.btn_select_folder = ctk.CTkButton(self.sidebar_frame, text="2. Select Video Folder", command=self.select_video_folder, **btn_params)
        self.btn_select_folder.grid(row=2, column=0, padx=20, pady=10)

        self.btn_select_bg = ctk.CTkButton(self.sidebar_frame, text="3. Select Background Image", command=self.select_background, **btn_params)
        self.btn_select_bg.grid(row=3, column=0, padx=20, pady=10)

        self.lbl_limit = ctk.CTkLabel(self.sidebar_frame, text="Batch Limit (Pause after X videos):", font=ctk.CTkFont(weight="bold"))
        self.lbl_limit.grid(row=4, column=0, padx=20, pady=(20, 0), sticky="w")
        
        self.entry_limit = ctk.CTkEntry(self.sidebar_frame, placeholder_text="e.g., 5 or 10", width=270)
        self.entry_limit.insert(0, "5") 
        self.entry_limit.grid(row=5, column=0, padx=20, pady=(5, 10))

        self.compress_var = ctk.IntVar(value=1)
        self.switch_compress = ctk.CTkSwitch(self.sidebar_frame, text="Enable Deep Compression", variable=self.compress_var, font=ctk.CTkFont(size=12, weight="bold"))
        self.switch_compress.grid(row=6, column=0, padx=20, pady=(15, 10), sticky="w")

        self.btn_settings = ctk.CTkButton(self.sidebar_frame, text="⚙️ SETTINGS & PREVIEW", fg_color="#333333", hover_color="#444444", command=self.open_settings_window, **btn_params)
        self.btn_settings.grid(row=8, column=0, padx=20, pady=(10, 10))

        self.btn_run = ctk.CTkButton(self.sidebar_frame, text="▶ START BATCH", fg_color="#28a745", hover_color="#218838", text_color="white", **btn_params)
        self.btn_run.configure(command=self.start_pipeline_thread)
        self.btn_run.grid(row=9, column=0, padx=20, pady=(10, 30))

        # --- MAIN AREA ---
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        self.main_frame.grid_rowconfigure(1, weight=1) 
        self.main_frame.grid_columnconfigure(0, weight=1)

        # Header Frame to hold both the Title and the Counter
        self.queue_header_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.queue_header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        self.queue_header_frame.grid_columnconfigure(0, weight=1)
        self.queue_header_frame.grid_columnconfigure(1, weight=1)

        self.queue_label = ctk.CTkLabel(self.queue_header_frame, text="Batch Queue (Sequential)", font=ctk.CTkFont(size=20, weight="bold"))
        self.queue_label.grid(row=0, column=0, sticky="w")

        # New Counter Label (Size 11 to match 'Resized: Yes', pushed to extreme right)
        self.queue_counter_label = ctk.CTkLabel(self.queue_header_frame, text="", font=ctk.CTkFont(size=11), text_color="#a3a3a3")
        self.queue_counter_label.grid(row=0, column=1, sticky="e", padx=(0, 5))

        self.queue_frame = ctk.CTkScrollableFrame(self.main_frame, fg_color="#1e1e1e", corner_radius=10)
        self.queue_frame.grid(row=1, column=0, sticky="nsew")

    # ==========================================
    # SETTINGS & LIVE PREVIEW WINDOW
    # ==========================================
    def open_settings_window(self):
        settings_win = ctk.CTkToplevel(self)
        settings_win.title("Settings & Live Preview")
        settings_win.geometry("900x550")
        settings_win.attributes('-topmost', True)
        settings_win.grab_set() 

        settings_win.grid_columnconfigure(0, weight=1)
        settings_win.grid_columnconfigure(1, weight=2)
        settings_win.grid_rowconfigure(0, weight=1)

        ctrl_frame = ctk.CTkFrame(settings_win, fg_color="transparent")
        ctrl_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        
        ctk.CTkLabel(ctrl_frame, text="Layout Controls", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w", pady=(0, 15))

        self.lbl_shrink_val = ctk.CTkLabel(ctrl_frame, text=f"Shrink Factor: {int(self.cfg['shrink']*100)}%")
        self.lbl_shrink_val.pack(anchor="w")
        self.slider_shrink = ctk.CTkSlider(ctrl_frame, from_=0.5, to=1.0, command=self.update_live_preview)
        self.slider_shrink.set(self.cfg['shrink'])
        self.slider_shrink.pack(fill="x", pady=(0, 20))

        self.lbl_bot_val = ctk.CTkLabel(ctrl_frame, text=f"Scan Area (Bottom): {int(self.cfg['bottom_pct']*100)}%")
        self.lbl_bot_val.pack(anchor="w")
        self.slider_bot = ctk.CTkSlider(ctrl_frame, from_=0.05, to=0.50, command=self.update_live_preview)
        self.slider_bot.set(self.cfg['bottom_pct'])
        self.slider_bot.pack(fill="x", pady=(0, 20))

        self.lbl_corn_val = ctk.CTkLabel(ctrl_frame, text=f"Corner Ignore Mask: {int(self.cfg['corner_pct']*100)}%")
        self.lbl_corn_val.pack(anchor="w")
        self.slider_corn = ctk.CTkSlider(ctrl_frame, from_=0.0, to=0.50, command=self.update_live_preview)
        self.slider_corn.set(self.cfg['corner_pct'])
        self.slider_corn.pack(fill="x", pady=(0, 20))

        ctk.CTkLabel(ctrl_frame, text="Compression Settings", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w", pady=(15, 15))
        
        self.lbl_crf_val = ctk.CTkLabel(ctrl_frame, text=f"CRF (Lower = Better Qty): {self.cfg['crf']}")
        self.lbl_crf_val.pack(anchor="w")
        self.slider_crf = ctk.CTkSlider(ctrl_frame, from_=18, to=35, number_of_steps=17, command=self.update_crf_label)
        self.slider_crf.set(self.cfg['crf'])
        self.slider_crf.pack(fill="x", pady=(0, 20))

        ctk.CTkLabel(ctrl_frame, text="Encoding Preset:").pack(anchor="w")
        self.combo_preset = ctk.CTkComboBox(ctrl_frame, values=["ultrafast", "superfast", "veryfast", "faster", "fast", "medium", "slow", "slower", "veryslow"])
        self.combo_preset.set(self.cfg['preset'])
        self.combo_preset.pack(fill="x", pady=(0, 20))

        btn_save = ctk.CTkButton(ctrl_frame, text="Save Settings", fg_color="#28a745", hover_color="#218838", command=lambda: self.save_settings(settings_win))
        btn_save.pack(fill="x", side="bottom", pady=10)

        preview_frame = ctk.CTkFrame(settings_win, fg_color="#1e1e1e", corner_radius=10)
        preview_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        
        self.res_selector = ctk.CTkSegmentedButton(
            preview_frame,
            values=["1920x1080 (16:9)", "1280x720 (16:9)", "1004x720 (Custom)"],
            command=self.update_live_preview
        )
        self.res_selector.set("1920x1080 (16:9)")
        self.res_selector.pack(pady=(10, 5))
        
        legend_frame = ctk.CTkFrame(preview_frame, fg_color="transparent")
        legend_frame.pack(pady=(0, 10))
        ctk.CTkLabel(legend_frame, text="■ Shrink Layout", text_color="#a855f7", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        ctk.CTkLabel(legend_frame, text="■ OCR Watch Zone", text_color="#3B8ED0", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)

        self.preview_lbl = ctk.CTkLabel(preview_frame, text="")
        self.preview_lbl.pack(expand=True, padx=20, pady=20)
        
        self.update_live_preview()

    def update_crf_label(self, val):
        self.lbl_crf_val.configure(text=f"CRF (Lower = Better Qty): {int(val)}")

    def update_live_preview(self, *args):
        shrink = self.slider_shrink.get()
        bot = self.slider_bot.get()
        corn = self.slider_corn.get()

        self.lbl_shrink_val.configure(text=f"Shrink Factor: {int(shrink*100)}%")
        self.lbl_bot_val.configure(text=f"Scan Area (Bottom): {int(bot*100)}%")
        self.lbl_corn_val.configure(text=f"Corner Ignore Mask: {int(corn*100)}%")

        # Dynamically calculate preview dimensions based on selected tab
        selected_res = getattr(self, 'res_selector', None)
        res_val = selected_res.get() if selected_res else "1920"

        if "1004" in res_val:
            w, h = 502, 360 # Locks height to 360 to prevent UI clipping, shrinks width
        else:
            w, h = 640, 360 # Standard 16:9 for 1920x1080 and 1280x720
        
        if self.bg_path and os.path.exists(self.bg_path):
            try:
                bg_img = Image.open(self.bg_path).convert("RGBA")
                img = ImageOps.fit(bg_img, (w, h), Image.Resampling.LANCZOS)
            except Exception:
                img = Image.new('RGBA', (w, h), color='#2b2b2b')
        else:
            img = Image.new('RGBA', (w, h), color='#2b2b2b')

        draw = ImageDraw.Draw(img, "RGBA")

        # Draw Shrink Box
        sw, sh = w * shrink, h * shrink
        sx0, sy0 = (w - sw) / 2, 0 
        sx1, sy1 = sx0 + sw, sy0 + sh
        draw.rectangle([sx0, sy0, sx1, sy1], outline='#a855f7', width=4)
        draw.text((sx0 + 10, sy0 + 10), "Video Boundaries", fill="#a855f7")

        # Draw OCR Watch Zone
        wy0 = h * (1 - bot)
        wx0 = w * corn
        wx1 = w * (1 - corn)
        draw.rectangle([wx0, wy0, wx1, h], fill=(59, 142, 208, 60), outline='#3B8ED0', width=3)
        draw.text((wx0 + 10, wy0 + 10), "OCR Watch Area", fill="#3B8ED0")

        ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(w, h))
        self.preview_lbl.configure(image=ctk_img)
        self.preview_lbl.image = ctk_img

    def save_settings(self, win):
        self.cfg["shrink"] = self.slider_shrink.get()
        self.cfg["bottom_pct"] = self.slider_bot.get()
        self.cfg["corner_pct"] = self.slider_corn.get()
        self.cfg["crf"] = int(self.slider_crf.get())
        self.cfg["preset"] = self.combo_preset.get()
        win.destroy()

    # ==========================================
    # FILE SELECTION & MAIN LOGIC
    # ==========================================
    def select_excel(self):
        self.excel_path = fd.askopenfilename(title='Select Excel Tracking File', filetypes=[('Excel files', '*.xlsx')])
        if self.excel_path: self.btn_select_excel.configure(text="1. Excel Linked ✓", fg_color="#3a5a40")

    def select_video_folder(self):
        self.video_folder = fd.askdirectory(title='Select Folder Containing All Videos')
        if self.video_folder:
            self.output_video_folder = os.path.join(self.video_folder, "Final_Render_Folder")
            if not os.path.exists(self.output_video_folder): os.makedirs(self.output_video_folder)
            self.btn_select_folder.configure(text="2. Folder Linked ✓", fg_color="#3a5a40")

    def select_background(self):
        self.bg_path = fd.askopenfilename(title='Select Background Image', filetypes=[('Image files', '*.jpg *.jpeg *.png')])
        if self.bg_path: self.btn_select_bg.configure(text="3. Background Linked ✓", fg_color="#3a5a40")

    def cancel_batch(self):
        self.cancel_flag = True
        self.btn_run.configure(state="disabled", text="CANCELLING...")

    def start_pipeline_thread(self):
        if not self.excel_path or not self.video_folder or not self.bg_path: return
        try:
            batch_limit = int(self.entry_limit.get())
            if batch_limit <= 0: raise ValueError
        except ValueError: return

        self.cancel_flag = False
        
        self.btn_run.configure(text="⏹ CANCEL BATCH", fg_color="#ff4c4c", hover_color="#cc0000", command=self.cancel_batch)
        self.update_idletasks()
        
        self.btn_select_excel.configure(state="disabled")
        self.btn_select_folder.configure(state="disabled")
        self.btn_select_bg.configure(state="disabled")
        self.entry_limit.configure(state="disabled")
        self.switch_compress.configure(state="disabled")
        self.btn_settings.configure(state="disabled")
        
        threading.Thread(target=self.run_master_controller, args=(batch_limit,), daemon=True).start()

    def ask_to_continue(self):
        self.user_choice = None
        self.wait_event.clear()
        self.after(0, self._create_popup)
        self.wait_event.wait() 
        return self.user_choice

    def _create_popup(self):
        popup = ctk.CTkToplevel(self)
        popup.title("Batch Complete")
        popup.geometry("350x180")
        popup.attributes('-topmost', True)
        popup.grab_set() 
        
        lbl = ctk.CTkLabel(popup, text="Current batch finished safely.\nWould you like to process the next batch?", font=ctk.CTkFont(size=14))
        lbl.pack(pady=30)
        
        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20)
        
        def on_yes():
            self.user_choice = True
            self.wait_event.set()
            popup.destroy()
            
        def on_no():
            self.user_choice = False
            self.wait_event.set()
            popup.destroy()

        ctk.CTkButton(btn_frame, text="Continue", fg_color="#28a745", hover_color="#218838", command=on_yes, width=120).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Stop Here", fg_color="#ff4c4c", hover_color="#cc0000", command=on_no, width=120).pack(side="right", padx=10)
        popup.protocol("WM_DELETE_WINDOW", on_no)


    def update_queue_label(self, completed, total):
        if total > 0:
            self.queue_counter_label.configure(text=f"{completed} / {total} Completed")
        else:
            self.queue_counter_label.configure(text="")

    def show_error_popup(self, message):
        popup = ctk.CTkToplevel(self)
        popup.title("Error: File Locked")
        popup.geometry("380x150")
        popup.attributes('-topmost', True)
        popup.grab_set()
        ctk.CTkLabel(popup, text=message, wraplength=340, font=ctk.CTkFont(size=14)).pack(pady=30)
        ctk.CTkButton(popup, text="OK", command=popup.destroy, width=120).pack()

    def setup_excel_headers(self, sheet, wb):
        # Requested exact order
        headers_to_add = [
            "Original Resolution", "Audio Status", "QA Status", 
            "Total Resizes", "Elapsed Time", "Original File Size", 
            "Compressed File Size", "Compression Used", 
            "Overlap Timestamps", "Obscured Text"
        ]
        
        max_col = 1
        while sheet.cell(row=1, column=max_col).value is not None: max_col += 1

        self.col_indices = {}
        # Map all existing headers (including "Duration" for the validation check)
        for col in range(1, max_col):
            val = str(sheet.cell(row=1, column=col).value).strip()
            self.col_indices[val] = col

        next_blank = max_col
        needs_save = False
        for header in headers_to_add:
            if header not in self.col_indices:
                sheet.cell(row=1, column=next_blank, value=header)
                self.col_indices[header] = next_blank
                next_blank += 1
                needs_save = True
        if needs_save: wb.save(self.excel_path)


    def run_master_controller(self, batch_limit):
        try:
            if os.path.basename(self.excel_path).startswith("~$"):
                return

            try:
                test_wb = openpyxl.load_workbook(self.excel_path)
                test_wb.save(self.excel_path)
            except PermissionError:
                self.after(0, lambda: self.show_error_popup("Your Excel file is currently open. Please close Microsoft Excel and try again."))
                return
            except Exception:
                return

            if not os.path.exists(self.output_video_folder):
                os.makedirs(self.output_video_folder, exist_ok=True)

            wb = openpyxl.load_workbook(self.excel_path)
            sheet = wb.active
            self.setup_excel_headers(sheet, wb)
            
            all_pending_jobs = []
            for row in range(2, sheet.max_row + 1):
                raw_title = sheet.cell(row=row, column=1).value
                if not raw_title: continue
                
                if str(raw_title).startswith("~$") or str(raw_title).startswith("."):
                    continue

                current_status = str(sheet.cell(row=row, column=self.col_indices.get("QA Status", 0)).value or "")
                if "Fixed" in current_status or current_status == "Clear" or "mismatch" in current_status: 
                    continue 
                
                # Fetch expected duration if the column exists
                expected_dur = None
                if "Duration" in self.col_indices:
                    expected_dur = sheet.cell(row=row, column=self.col_indices["Duration"]).value
                    
                all_pending_jobs.append({"row": row, "title": raw_title, "expected_duration": expected_dur})

            total_jobs = len(all_pending_jobs)
            if total_jobs == 0: 
                self.after(0, lambda: self.update_queue_label(0, 0))
                return

            completed_jobs = 0
            self.after(0, lambda: self.update_queue_label(completed_jobs, total_jobs))

            for i in range(0, total_jobs, batch_limit):
                if self.cancel_flag: break

                current_batch = all_pending_jobs[i : i + batch_limit]
                for widget in self.queue_frame.winfo_children(): widget.destroy()
                
                job_tiles = {}
                for job in current_batch:
                    tile = JobTile(self.queue_frame, video_name=job['title'])
                    tile.pack(fill="x", pady=5)
                    job_tiles[job['row']] = tile

                for job in current_batch:
                    if self.cancel_flag: break
                    self.process_single_video(job, job_tiles[job['row']])
                    completed_jobs += 1
                    self.after(0, lambda c=completed_jobs, t=total_jobs: self.update_queue_label(c, t))
                
                if self.cancel_flag: break

                if i + batch_limit < total_jobs:
                    if not self.ask_to_continue(): break

        except Exception as e:
            print(f"Master Error: {str(e)}")
        finally:
            self.btn_run.configure(state="normal", text="▶ START BATCH", fg_color="#28a745", hover_color="#218838")
            self.btn_select_excel.configure(state="normal", fg_color=["#3B8ED0", "#1F6AA5"])
            self.btn_select_folder.configure(state="normal", fg_color=["#3B8ED0", "#1F6AA5"])
            self.btn_select_bg.configure(state="normal", fg_color=["#3B8ED0", "#1F6AA5"])
            self.entry_limit.configure(state="normal")
            self.switch_compress.configure(state="normal")
            self.btn_settings.configure(state="normal")

    def process_single_video(self, job, tile):
        if self.cancel_flag: 
            tile.update_tile("Cancelled", 1.0, color="#ff4c4c", stage="Aborted", resized_status="N/A")
            return

        row = job['row']
        raw_title = job['title']
        expected_duration_str = job.get('expected_duration')
        
        tile.start_timer()
        tile.update_tile("Locating file...", 0.05, "#ffcc00", stage="Initialization")
        video_path = self.find_video_file(raw_title)

        # UPDATED: Returns a native time object so Excel can calculate/sum it
        def get_elapsed():
            if not tile.start_time: return datetime.time(0, 0, 0)
            elapsed_sec = int(time.time() - tile.start_time)
            hours, remainder = divmod(elapsed_sec, 3600)
            mins, secs = divmod(remainder, 60)
            return datetime.time(hour=hours, minute=mins, second=secs)

        def get_mb(path):
            return f"{os.path.getsize(path) / (1024*1024):.1f} MB" if os.path.exists(path) else "N/A"

        if not video_path:
            self.save_to_excel(row, qa_status="File Not Found", elapsed_time=get_elapsed(), total_resizes=0)
            tile.stop_timer()
            tile.update_tile("File Not Found", 1.0, color="#ff4c4c", stage="Halted", resized_status="Error")
            return

        # Fetch Video Metadata
        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS) or 30
        frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
        vid_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        vid_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        resolution_str = f"{vid_w}x{vid_h}"
        actual_duration_sec = (frames / fps) if fps > 0 else 0
        cap.release()

        # Duration Match Verification (2-second tolerance)
        if expected_duration_str:
            try:
                parts = str(expected_duration_str).split(":")
                if len(parts) == 2:
                    expected_sec = int(parts[0]) * 60 + int(parts[1])
                elif len(parts) == 3:
                    expected_sec = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                else:
                    expected_sec = actual_duration_sec
                
                if abs(expected_sec - actual_duration_sec) > 2.0:
                    self.save_to_excel(row, qa_status="Duration mismatch / Wrong video", resolution=resolution_str, elapsed_time=get_elapsed(), total_resizes=0)
                    tile.stop_timer()
                    tile.update_tile("Wrong Video / Duration", 1.0, color="#ff4c4c", stage="Halted", resized_status="Error")
                    return
            except Exception:
                pass # Proceed safely if Excel duration format is unreadable

        if fps > 0 and frames > 0:
            tile.set_duration(actual_duration_sec)

        orig_file_size = get_mb(video_path)
        video_filename = os.path.basename(video_path)
        tile.update_tile("Checking Audio...", 0.1, "#ffcc00", stage="Audio Check")
        audio_status = self.check_audio(video_path)

        intervals = self.scan_video(video_path, tile)
        
        if getattr(self, 'cancel_flag', False):
            tile.stop_timer()
            tile.update_tile("Cancelled during scan", 1.0, color="#ff4c4c", stage="Aborted", resized_status="N/A")
            return

        if intervals is None:
            self.save_to_excel(row, qa_status="Error Reading Video", resolution=resolution_str, elapsed_time=get_elapsed(), total_resizes=0)
            tile.stop_timer()
            tile.update_tile("Read Error", 1.0, color="#ff4c4c", stage="Halted", resized_status="Error")
            return
        elif len(intervals) == 0:
            tile.update_tile("Copying Clean Video...", 0.9, "#3B8ED0", stage="File Transfer")
            out_path = os.path.join(self.output_video_folder, video_filename)
            shutil.copy2(video_path, out_path)
            self.save_to_excel(row, qa_status="Clear", resolution=resolution_str, audio_status=audio_status, timestamps="None", text_data="None", elapsed_time=get_elapsed(), total_resizes=0, orig_size=orig_file_size, comp_size=get_mb(out_path), comp_used="Direct Copy")
            tile.stop_timer()
            tile.update_tile("Clear & Copied", 1.0, color="#28a745", stage="Complete", resized_status="No (Original Copied)")
            return
            
        all_timestamps = ", ".join([f"{self.format_timestamp(i['start'])}-{self.format_timestamp(i['end'])}" for i in intervals])
        all_texts = " | ".join([text for i in intervals for text in i['texts']])
        num_resizes = len(intervals)

        success = self.render_and_compress(video_path, intervals, tile)
        
        base_name, ext = os.path.splitext(video_filename)
        final_output_path = os.path.join(self.output_video_folder, f"{base_name}_resized{ext}")
        comp_file_size = get_mb(final_output_path)
        
        is_compressed = self.compress_var.get() == 1
        compression_used = f"{self.cfg['preset']}_{self.cfg['crf']}" if is_compressed else "Hardware Render (Fast)"

        if success:
            success_msg = "Fixed & Compressed" if is_compressed else "Fixed without compression"
            self.save_to_excel(row, qa_status=success_msg, resolution=resolution_str, audio_status=audio_status, timestamps=all_timestamps, text_data=all_texts, elapsed_time=get_elapsed(), total_resizes=num_resizes, orig_size=orig_file_size, comp_size=comp_file_size, comp_used=compression_used)
            tile.stop_timer()
            tile.update_tile(success_msg, 1.0, color="#28a745", stage="Complete", resized_status="Yes")
        else:
            self.save_to_excel(row, qa_status="Render Error", resolution=resolution_str, audio_status=audio_status, elapsed_time=get_elapsed(), total_resizes=num_resizes, orig_size=orig_file_size)
            tile.stop_timer()
            tile.update_tile("Render Failed", 1.0, color="#ff4c4c", stage="Halted", resized_status="Error")

    def save_to_excel(self, row, qa_status, resolution="N/A", audio_status="N/A", timestamps=None, text_data=None, elapsed_time=None, total_resizes=None, orig_size=None, comp_size=None, comp_used=None):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            sheet = wb.active
            
            def write_if_exists(header, value):
                if header in self.col_indices and value is not None:
                    sheet.cell(row=row, column=self.col_indices[header], value=value)
                    
            write_if_exists("QA Status", qa_status)
            write_if_exists("Original Resolution", resolution)
            write_if_exists("Audio Status", audio_status)
            write_if_exists("Overlap Timestamps", timestamps)
            write_if_exists("Obscured Text", text_data)
            write_if_exists("Elapsed Time", elapsed_time)
            write_if_exists("Total Resizes", total_resizes)
            write_if_exists("Original File Size", orig_size)
            write_if_exists("Compressed File Size", comp_size)
            write_if_exists("Compression Used", comp_used)
            
            wb.save(self.excel_path)
        except Exception as e: 
            print(f"Excel Save Error: {e}")

    def normalize_string(self, s):
        if not s: return ""
        return re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()

    def find_video_file(self, excel_title):
        normalized_title = self.normalize_string(excel_title).replace('mp4', '')
        for filename in os.listdir(self.video_folder):
            if filename.lower().endswith(('.mp4', '.mov', '.avi')):
                normalized_filename = self.normalize_string(filename).replace('mp4', '').replace('mov', '').replace('avi', '')
                if normalized_title == normalized_filename or normalized_title in normalized_filename: return os.path.join(self.video_folder, filename)
        return None

    def check_audio(self, video_path):
        try:
            cmd = ["ffprobe", "-v", "error", "-show_entries", "stream=codec_type", "-of", "csv=p=0", video_path]
            return "Has Audio" if "audio" in subprocess.check_output(cmd, text=True).strip() else "No Audio"
        except Exception: return "Audio Check Error"

    def format_timestamp(self, seconds):
        mins, secs = divmod(int(seconds), 60)
        return f"{mins:02d}:{secs:02d}"

    def scan_video(self, video_path, tile):
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened(): return None

        fps = cap.get(cv2.CAP_PROP_FPS) or 30
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        
        bot_pct = self.cfg["bottom_pct"]
        corn_pct = self.cfg["corner_pct"]
        frame_interval = int(fps * self.cfg.get("CHECK_INTERVAL", 1.0))
        
        frame_count, current_start = 0, 0
        raw_intervals, current_texts = [], []
        is_overlapping = False

        while cap.isOpened():
            if getattr(self, 'cancel_flag', False):
                cap.release()
                return None
                
            ret, frame = cap.read()
            if not ret: break

            # Scan the frame once per second
            if frame_count % frame_interval == 0:
                if total_frames > 0:
                    prog = 0.1 + (0.5 * (frame_count / total_frames))
                    tile.update_tile(f"Scanning ({self.format_timestamp(frame_count/fps)})...", prog, "#3B8ED0", stage="OCR Analysis")

                h, w, _ = frame.shape
                
                # Isolate the exact blue box based on your UI sliders
                subtitle_zone = frame[int(h * (1 - bot_pct)):h, 0:w]
                gray_zone = cv2.cvtColor(subtitle_zone, cv2.COLOR_BGR2GRAY)
                mask_w = int(w * corn_pct)
                gray_zone[:, :mask_w] = 255           
                gray_zone[:, w - mask_w:] = 255   
                
                # Standard blur and threshold
                blurred_zone = cv2.medianBlur(gray_zone, 3)
                _, thresh_zone = cv2.threshold(blurred_zone, 150, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
                
                # Run OCR
                ocr_data = pytesseract.image_to_data(thresh_zone, output_type=pytesseract.Output.DICT)
                
                valid_word_count = 0
                n_boxes = len(ocr_data['text'])
                
                for idx in range(n_boxes):
                    try:
                        conf = float(ocr_data['conf'][idx])
                    except ValueError:
                        continue
                        
                    word = ocr_data['text'][idx].strip()
                    word_clean = re.sub(r'[^a-zA-Z0-9]', '', word)
                    
                    if "kaplan" in word.lower():
                        continue
                        
                    # SIMPLE RULE: Is it a real word (3+ letters) with decent confidence (>60%)?
                    if len(word_clean) >= 3 and conf >= 60.0:
                        valid_word_count += 1
                        if word not in current_texts: 
                            current_texts.append(word)

                text_present = (valid_word_count > 0)

                # Track the exact start and end times of the text
                if text_present and not is_overlapping:
                    is_overlapping = True
                    current_start = frame_count / fps
                elif not text_present and is_overlapping:
                    is_overlapping = False
                    raw_intervals.append({'start': current_start, 'end': frame_count / fps, 'texts': current_texts})
                    current_texts = [] 

            frame_count += 1

        if is_overlapping: 
            raw_intervals.append({'start': current_start, 'end': (frame_count / fps), 'texts': current_texts})
        cap.release()
        
        if not raw_intervals: return []

        # ==========================================
        # 15-SECOND BRIDGE LOGIC
        # ==========================================
        raw_intervals.sort(key=lambda x: x['start'])
        smoothed_intervals = []
        
        current_interval = raw_intervals[0]
        current_interval['start'] = max(0, current_interval['start'] - self.cfg["time_buffer_start"])

        for next_interval in raw_intervals[1:]:
            next_start_buffered = max(0, next_interval['start'] - self.cfg["time_buffer_start"])
            gap = next_start_buffered - current_interval['end']
            
            # If the gap between text overlaps is 15 seconds or less, combine them into one long shrink
            if gap <= 15.0:
                current_interval['end'] = next_interval['end']
                current_interval['texts'] = list(set(current_interval['texts'] + next_interval['texts']))
            else:
                smoothed_intervals.append(current_interval)
                current_interval = next_interval
                current_interval['start'] = next_start_buffered

        smoothed_intervals.append(current_interval)
        
        # ==========================================
        # ANTI-JERK FILTER: OMIT SHRINKS UNDER 2 SECONDS
        # ==========================================
        final_intervals = []
        for interval in smoothed_intervals:
            duration = interval['end'] - interval['start']
            
            # Only keep the shrink if it lasts for 2.0 seconds or longer
            if duration >= 3.0:
                final_intervals.append(interval)
            else:
                print(f"Skipping micro-shrink at {interval['start']} (Duration: {duration}s)")

        return final_intervals
    
    def render_and_compress(self, input_path, intervals, tile):
        if getattr(self, 'cancel_flag', False):
            return False

        if not os.path.exists(self.output_video_folder):
            os.makedirs(self.output_video_folder, exist_ok=True)

        video_filename = os.path.basename(input_path)
        base_name, ext = os.path.splitext(video_filename)
        
        rendered_temp = os.path.join(self.output_video_folder, f"TEMP_{video_filename}")
        final_output = os.path.join(self.output_video_folder, f"{base_name}_resized{ext}")

        cap = cv2.VideoCapture(input_path)
        vid_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        vid_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        cap.release()

        bg_input_source = self.bg_path
        if not self.bg_path or not os.path.exists(self.bg_path):
            fallback_bg = os.path.join(self.output_video_folder, "SAFE_FALLBACK_BG.png")
            try:
                # Updated to dynamically draw the fallback image to the native resolution
                safe_img = Image.new('RGB', (vid_w, vid_h), color='#1e1e2f')
                safe_img.save(fallback_bg)
                bg_input_source = fallback_bg
            except Exception:
                bg_input_source = f"color=c=black:s={vid_w}x{vid_h}"

        shrink = self.cfg["shrink"]
        enable_expr = "+".join([f"between(t,{i['start']},{i['end']})" for i in intervals])
        
        # Updated to inject the dynamic vid_w and vid_h instead of hardcoding 1920:1080
        filter_complex = (
            f"[1:v]scale={vid_w}:{vid_h}:force_original_aspect_ratio=increase,crop={vid_w}:{vid_h}[bg];"
            f"[0:v]scale={vid_w}:{vid_h}:force_original_aspect_ratio=decrease,pad={vid_w}:{vid_h}:(ow-iw)/2:(oh-ih)/2:color=black[orig_standard];"
            f"[orig_standard]split=2[orig][to_shrink];"
            f"[to_shrink]scale=iw*{shrink}:ih*{shrink}[shrunk];"
            f"[bg][shrunk]overlay=(W-w)/2:0[padded];"
            f"[orig][padded]overlay=enable='{enable_expr}'[v]"
        )

        try:
            do_compression = self.compress_var.get() == 1

            if do_compression:
                tile.update_tile("Building layout...", 0.65, "#a855f7", stage="Hardware Render")
                
                bg_args = ["-i", bg_input_source] if "color=" not in bg_input_source else ["-f", "lavfi", "-i", bg_input_source]
                
                cmd = ["ffmpeg", "-y", "-i", input_path, "-loop", "1", "-t", "1"] + bg_args + [
                    "-filter_complex", filter_complex, "-map", "[v]", "-map", "0:a?",
                    "-c:v", "h264_videotoolbox", "-b:v", "4M", "-c:a", "aac", rendered_temp
                ]
                subprocess.run(cmd, check=True, capture_output=True)

                if self.cancel_flag: return False 

                tile.update_tile("Crunching file size...", 0.85, "#a855f7", stage="Software Compression")
                subprocess.run([
                    "ffmpeg", "-y", "-i", rendered_temp, "-c:v", "libx264", 
                    "-crf", str(self.cfg["crf"]), 
                    "-preset", self.cfg["preset"], 
                    "-c:a", "copy", final_output
                ], check=True, capture_output=True)

                if os.path.exists(rendered_temp): os.remove(rendered_temp)
            else:
                tile.update_tile("Fast Hardware Render...", 0.75, "#a855f7", stage="Fast Render")
                bg_args = ["-i", bg_input_source] if "color=" not in bg_input_source else ["-f", "lavfi", "-i", bg_input_source]
                
                cmd = ["ffmpeg", "-y", "-i", input_path, "-loop", "1", "-t", "1"] + bg_args + [
                    "-filter_complex", filter_complex, "-map", "[v]", "-map", "0:a?",
                    "-c:v", "h264_videotoolbox", "-b:v", "4M", "-c:a", "aac", final_output
                ]
                subprocess.run(cmd, check=True, capture_output=True)

            return True
        except subprocess.CalledProcessError as ffmpeg_error:
            print(f"FFmpeg Rendering Crash Error: {ffmpeg_error.stderr.decode('utf-8', errors='ignore')}")
            return False

if __name__ == "__main__":
    app = VideoResizerApp()
    app.mainloop()